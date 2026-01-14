import streamlit as st
import pandas as pd
import numpy as np
import io
import re
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

st.set_page_config(page_title="IPQC CPK & Yield", layout="wide")
st.title("📊 IPQC CPK & Yield 報表生成器")

# --- 1. 定義站點順序 ---
TARGET_ORDER = [
    "MLA assy installation", "Mirror attachment", "Barrel attachment",
    "Condenser lens attach", "LED Module  attachment",
    "ILLU Module cover attachment", "Relay lens attachment",
    "LED FLEX GRAPHITE-1", "reflector attach", "singlet attach",
    "HWP Mylar attach", "PBS attachment", "Doublet attachment",
    "Top cover installation", "PANEL PRECISION AA（LAA）",
    "POST DAA INSPECTION", "PANEL FLEX ASSY",
    "LCOS GRAPHITE ATTACH", "DE OQC"
]

def normalize_name(name):
    return str(name).lower().replace(" ", "").replace("（", "").replace("）", "").replace("(", "").replace(")", "").replace("-", "").replace("_", "")

TARGET_MAP = {normalize_name(n): n for n in TARGET_ORDER}

# --- 2. CPK 計算函式 ---
def calculate_cpk(data, usl, lsl):
    data = pd.to_numeric(data, errors="coerce").dropna()
    if len(data) < 2:
        return np.nan
    std = data.std(ddof=1)
    if std == 0:
        return np.nan
    mean = data.mean()
    cpu = (usl - mean) / (3 * std) if not pd.isna(usl) else np.nan
    cpl = (mean - lsl) / (3 * std) if not pd.isna(lsl) else np.nan
    
    if not pd.isna(cpu) and not pd.isna(cpl):
        return min(cpu, cpl)
    elif not pd.isna(cpu):
        return cpu
    elif not pd.isna(cpl):
        return cpl
    else:
        return np.nan

# --- 3. 主程式 ---
uploaded = st.file_uploader("📂 上傳 IPQC Excel (.xlsx)", type=["xlsx"])

if uploaded:
    try:
        xls = pd.ExcelFile(uploaded)
        yield_rows, cpk_rows = [], []

        # 顯示進度條
        progress_bar = st.progress(0)
        sheet_count = len(xls.sheet_names)

        for i, sheet in enumerate(xls.sheet_names):
            # 更新進度
            progress_bar.progress((i + 1) / sheet_count)
            
            # 站點名稱比對
            norm = normalize_name(sheet)
            station = next((v for k, v in TARGET_MAP.items() if k in norm), None)
            
            # 特殊修正 (避免模糊比對錯誤)
            if "postdaa" in norm: station = "POST DAA INSPECTION"
            
            if not station:
                continue

            # 讀取 Sheet (不含 header，方便後續定位)
            df = pd.read_excel(uploaded, sheet_name=sheet, header=None)

            # 尋找關鍵字所在的列 (Row Index)
            def find_row(key):
                for r in range(min(80, len(df))):
                    row_str = " ".join(df.iloc[r].astype(str).str.lower())
                    if key in row_str:
                        return r
                return -1

            dim_row = find_row("dim")
            usl_row = find_row("usl")
            lsl_row = find_row("lsl")

            # ---------- Yield 計算 (維持原樣) ----------
            best_col, max_cnt = -1, 0
            # 掃描前 30 欄
            for c in range(min(30, df.shape[1])):
                col = df.iloc[:, c].astype(str).str.upper()
                cnt = (col == "OK").sum() + (col == "NG").sum()
                if cnt > max_cnt:
                    max_cnt, best_col = cnt, c

            if best_col != -1 and max_cnt > 0:
                col_data = df.iloc[:, best_col].astype(str).str.upper()
                ok = (col_data == "OK").sum()
                ng = (col_data == "NG").sum()
                yield_rows.append({
                    "Station": station,
                    "Total Qty": ok + ng,
                    "OK Qty": ok,
                    "NG Qty": ng,
                    "Yield": ok / (ok + ng) if (ok + ng) > 0 else 0
                })

            # ---------- CPK 計算 (新增 Config) ----------
            if dim_row == -1:
                continue

            # 解析 Dim No
            dim_headers = df.iloc[dim_row]
            dims = {
                i: str(v).strip() for i, v in enumerate(dim_headers)
                if str(v).strip() and str(v).lower() not in ["dim", "dimension", "nan", "date", "time", "no.", "remark"]
            }

            # 解析 USL / LSL
            usls, lsls = {}, {}
            if usl_row != -1:
                for idx, v in enumerate(df.iloc[usl_row]):
                    try: usls[idx] = float(v)
                    except: pass
            if lsl_row != -1:
                for idx, v in enumerate(df.iloc[lsl_row]):
                    try: lsls[idx] = float(v)
                    except: pass

            # 提取數據區塊
            start = max(dim_row, usl_row, lsl_row) + 1
            data = df.iloc[start:].copy()

            # 尋找 Date 欄位 (YYYY-MM-DD)
            date_col = -1
            for c in range(min(20, data.shape[1])):
                if data.iloc[:, c].astype(str).str.contains(r"202\d-\d{2}-\d{2}").any():
                    date_col = c
                    break
            
            if date_col != -1:
                # 提取 Date
                data["Date"] = data.iloc[:, date_col].astype(str).str.extract(r"(202\d-\d{2}-\d{2})")[0]
                
                # [新增] 提取 Config (預設抓取第 0 欄，通常是 SH2, SH3...)
                # 如果第 0 欄是 Date，則抓 Date+1
                config_col = 0 if date_col != 0 else 1
                data["Config"] = data.iloc[:, config_col].astype(str).fillna("")
                
                # 排除無效日期
                data = data.dropna(subset=["Date"])

                # 依 Date + Config 分組計算
                for (date_val, config_val), g in data.groupby(["Date", "Config"]):
                    for idx, dim_no in dims.items():
                        vals = g.iloc[:, idx]
                        n = pd.to_numeric(vals, errors="coerce").dropna().size
                        
                        if n > 1:
                            cpk = calculate_cpk(vals, usls.get(idx), lsls.get(idx))
                            
                            cpk_rows.append({
                                "Station": station,
                                "Dim No": dim_no,
                                "config": config_val,  # 新增欄位
                                "Date": date_val,
                                "Sample Size": n,
                                "USL": usls.get(idx, ""),
                                "LSL": lsls.get(idx, ""),
                                "CPK": cpk
                            })

        # ---------- 建立 DataFrames ----------
        df_yield = pd.DataFrame(yield_rows)
        df_cpk = pd.DataFrame(cpk_rows)

        # 排序
        if not df_yield.empty:
            df_yield['Station'] = pd.Categorical(df_yield['Station'], categories=TARGET_ORDER, ordered=True)
            df_yield = df_yield.sort_values('Station')
            
        if not df_cpk.empty:
            df_cpk['Station'] = pd.Categorical(df_cpk['Station'], categories=TARGET_ORDER, ordered=True)
            # 指定欄位順序: Station, Dim No, config, Date, ...
            cols = ["Station", "Dim No", "config", "Date", "Sample Size", "USL", "LSL", "CPK"]
            df_cpk = df_cpk[cols].sort_values(by=['Station', 'Dim No', 'Date'])

        # ---------- 顯示結果 ----------
        st.success("✅ 計算完成！")

        st.subheader("📈 Yield Summary")
        if not df_yield.empty:
            st.dataframe(df_yield.assign(Yield=lambda d: (d["Yield"] * 100).round(2).astype(str) + "%"), use_container_width=True)
        else:
            st.warning("未偵測到良率數據")

        st.subheader("📉 CPK Detail (CPK < 1.33 標示紅色)")
        if not df_cpk.empty:
            st.dataframe(
                df_cpk.style.applymap(
                    lambda v: "background-color:#ff9999" if isinstance(v, (int, float)) and v < 1.33 else "",
                    subset=["CPK"]
                ),
                use_container_width=True
            )
        else:
            st.warning("未偵測到 CPK 數據")

        # ---------- Excel 匯出 (Openpyxl) ----------
        output = io.BytesIO()
        with pd.ExcelWriter(output, engine="openpyxl") as writer:
            if not df_yield.empty:
                df_yield.to_excel(writer, sheet_name="Yield Summary", index=False)
            if not df_cpk.empty:
                df_cpk.to_excel(writer, sheet_name="CPK Detail", index=False)

        # 讀取剛寫入的 Excel 進行格式化 (標紅字)
        if not df_cpk.empty:
            wb = load_workbook(output)
            ws = wb["CPK Detail"]
            red_fill = PatternFill("solid", fgColor="FF6666")
            
            # 找出 CPK 所在的欄位索引 (1-based)
            header = [cell.value for cell in ws[1]]
            try:
                cpk_col_idx = header.index("CPK") + 1
                for r in range(2, ws.max_row + 1):
                    val = ws.cell(row=r, column=cpk_col_idx).value
                    if isinstance(val, (int, float)) and val < 1.33:
                        ws.cell(row=r, column=cpk_col_idx).fill = red_fill
            except ValueError:
                pass
            
            # 儲存格式化後的檔案
            final_output = io.BytesIO()
            wb.save(final_output)
            final_output.seek(0)
            data_to_download = final_output
        else:
            output.seek(0)
            data_to_download = output

        st.download_button(
            label="📥 下載完整報表 (Excel)",
            data=data_to_download,
            file_name="IPQC_Final_Report.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )

    except Exception as e:
        st.error(f"發生錯誤: {e}")